#!/usr/bin/env python3
"""
Update the "Result I want" sheet with ALL columns the asset uses (both non-custom and custom)
"""

import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill


def main():
    wb = openpyxl.load_workbook('D:/rr_repo/manual/everything_columnmapping.xlsx')

    # Styles
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_font_white = Font(bold=True, color='FFFFFF')
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)

    # Remove existing 'Result I want' sheet
    if 'Result I want' in wb.sheetnames:
        del wb['Result I want']

    ws = wb.create_sheet(title='Result I want')

    # All columns with their metadata
    all_columns = [
        {
            'name': 'Subject',
            'cp_slot': None,
            'description': 'Asset classification (primary categorization)',
            'enrichment': 'AI-determined leaf subject phrase, prefixed with AssetType/'
        },
        {
            'name': 'Title',
            'cp_slot': None,
            'description': 'Model name/designer name',
            'enrichment': 'Extracted from filename or AI vision analysis'
        },
        {
            'name': 'Company',
            'cp_slot': None,
            'description': 'Brand/designer or secondary identifier',
            'enrichment': 'Extracted from filename context or domain knowledge'
        },
        {
            'name': 'Album',
            'cp_slot': None,
            'description': 'Collection name',
            'enrichment': 'Extracted from filename or grouping metadata'
        },
        {
            'name': 'Author',
            'cp_slot': None,
            'description': 'Vendor/company name or source',
            'enrichment': 'Extracted from source origin metadata'
        },
        {
            'name': 'Period',
            'cp_slot': None,
            'description': 'Style, era, or temporal classification',
            'enrichment': 'AI vision analysis identifies design style'
        },
        {
            'name': 'Color',
            'cp_slot': 0,
            'description': 'Primary color/material',
            'enrichment': 'AI vision analysis extracts dominant color'
        },
        {
            'name': 'Location',
            'cp_slot': 1,
            'description': 'Usage context/location/habitat',
            'enrichment': 'AI reasoning determines appropriate usage context'
        },
        {
            'name': 'Form',
            'cp_slot': 2,
            'description': 'Shape/form/physical configuration',
            'enrichment': 'AI vision analysis identifies geometric form'
        },
        {
            'name': 'ChineseName',
            'cp_slot': 3,
            'description': 'Chinese scientific name (vegetation)',
            'enrichment': 'AI plant classification provides Chinese name'
        },
        {
            'name': 'LatinName',
            'cp_slot': 4,
            'description': 'Latin/Scientific name (vegetation)',
            'enrichment': 'AI plant classification provides scientific name'
        },
        {
            'name': 'Size',
            'cp_slot': 5,
            'description': 'Dimensions or scale classification',
            'enrichment': 'Extracted from metadata or AI scale estimation'
        },
        {
            'name': 'Code',
            'cp_slot': 6,
            'description': 'Material or reference code identifier',
            'enrichment': 'OCR from schedule documents'
        },
        {
            'name': 'Finish',
            'cp_slot': 7,
            'description': 'Surface finish or treatment',
            'enrichment': 'OCR from schedule documents + AI vision'
        },
    ]

    # Exact usage from the matrix in everything_columnmapping.md
    # True means the asset uses this column
    asset_usage = {
        'Furniture': {
            'Subject': True, 'Title': True, 'Company': True, 'Album': True,
            'Author': True, 'Period': True, 'Color': True, 'Location': True,
            'Form': True, 'ChineseName': False, 'LatinName': False,
            'Size': True, 'Code': False, 'Finish': False
        },
        'Fixture': {
            'Subject': True, 'Title': True, 'Company': True, 'Album': True,
            'Author': True, 'Period': True, 'Color': True, 'Location': True,
            'Form': True, 'ChineseName': False, 'LatinName': False,
            'Size': True, 'Code': False, 'Finish': False
        },
        'Vegetation': {
            'Subject': True, 'Title': True, 'Company': True, 'Album': False,
            'Author': False, 'Period': False, 'Color': True, 'Location': False,
            'Form': True, 'ChineseName': True, 'LatinName': True,
            'Size': True, 'Code': False, 'Finish': False
        },
        'Object': {
            'Subject': True, 'Title': True, 'Company': True, 'Album': True,
            'Author': True, 'Period': True, 'Color': True, 'Location': True,
            'Form': True, 'ChineseName': False, 'LatinName': False,
            'Size': True, 'Code': False, 'Finish': False
        },
        'Material': {
            'Subject': True, 'Title': True, 'Company': True, 'Album': False,
            'Author': False, 'Period': True, 'Color': True, 'Location': False,
            'Form': True, 'ChineseName': False, 'LatinName': False,
            'Size': True, 'Code': False, 'Finish': False
        },
        'Vehicle': {
            'Subject': True, 'Title': True, 'Company': True, 'Album': False,
            'Author': False, 'Period': True, 'Color': True, 'Location': False,
            'Form': False, 'ChineseName': False, 'LatinName': False,
            'Size': True, 'Code': False, 'Finish': False
        },
        'Layouts': {
            'Subject': True, 'Title': False, 'Company': False, 'Album': True,
            'Author': False, 'Period': False, 'Color': False, 'Location': True,
            'Form': True, 'ChineseName': False, 'LatinName': False,
            'Size': True, 'Code': False, 'Finish': False
        },
        'People': {
            'Subject': True, 'Title': False, 'Company': False, 'Album': False,
            'Author': False, 'Period': True, 'Color': True, 'Location': True,
            'Form': False, 'ChineseName': False, 'LatinName': False,
            'Size': False, 'Code': False, 'Finish': False
        },
        'VFX': {
            'Subject': True, 'Title': False, 'Company': False, 'Album': False,
            'Author': False, 'Period': False, 'Color': False, 'Location': False,
            'Form': False, 'ChineseName': False, 'LatinName': False,
            'Size': False, 'Code': False, 'Finish': False
        },
        'Schedule': {
            'Subject': True, 'Title': True, 'Company': True, 'Album': False,
            'Author': True, 'Period': False, 'Color': True, 'Location': True,
            'Form': False, 'ChineseName': False, 'LatinName': False,
            'Size': True, 'Code': True, 'Finish': True
        },
    }

    # Column A header is the row type, then CP_0 to CP_9 in columns B-K
    headers_cp = ['Row Type', 'CP_0', 'CP_1', 'CP_2', 'CP_3', 'CP_4', 'CP_5', 'CP_6', 'CP_7', 'CP_8', 'CP_9']

    current_row = 1

    # Process each asset type
    for asset_type, usage_map in asset_usage.items():
        # Asset type title row (spans all columns, bold)
        cell = ws.cell(row=current_row, column=1, value=asset_type)
        cell.font = Font(bold=True, size=14)
        current_row += 1

        # Column Name row
        for col_idx, header in enumerate(headers_cp, start=1):
            cell = ws.cell(row=current_row, column=col_idx, value=header)
            cell.font = header_font_white
            cell.fill = header_fill
            cell.border = thin_border
            cell.alignment = center_align
        current_row += 1

        # Description row
        col_idx_start = current_row
        for col_idx, header in enumerate(headers_cp, start=1):
            cell = ws.cell(row=current_row, column=col_idx)
            cell.border = thin_border
        cell = ws.cell(row=current_row, column=1, value='Description')
        cell.font = Font(bold=True)
        current_row += 1

        # Enrichment Method row
        for col_idx, header in enumerate(headers_cp, start=1):
            cell = ws.cell(row=current_row, column=col_idx)
            cell.border = thin_border
        cell = ws.cell(row=current_row, column=1, value='Enrichment Method')
        cell.font = Font(bold=True)
        current_row += 1

        # Now fill ALL used columns
        for col_info in all_columns:
            if usage_map[col_info['name']]:
                if col_info['cp_slot'] is None:
                    # Non-CP column goes to column A (Row Type column)
                    # We already have "Description" and "Enrichment Method" in column A, fill the content
                    desc_row = col_idx_start
                    enrich_row = col_idx_start + 1
                    ws.cell(row=desc_row, column=1, value=col_info['description']).border = thin_border
                    ws.cell(row=desc_row, column=1).alignment = left_align
                    ws.cell(row=enrich_row, column=1, value=col_info['enrichment']).border = thin_border
                    ws.cell(row=enrich_row, column=1).alignment = left_align
                    # Put column name in Column Name row (first column)
                    ws.cell(row=col_idx_start - 1, column=1, value=col_info['name']).border = thin_border
                    ws.cell(row=col_idx_start - 1, column=1).alignment = center_align
                else:
                    # Custom property - goes to corresponding CP column (B-K = CP_0-CP_9)
                    col_index = col_info['cp_slot'] + 2  # CP_0 is column 2 (B)
                    desc_row = col_idx_start
                    enrich_row = col_idx_start + 1
                    # Fill column name
                    ws.cell(row=col_idx_start - 1, column=col_index, value=col_info['name']).border = thin_border
                    ws.cell(row=col_idx_start - 1, column=col_index).alignment = center_align
                    # Fill description
                    ws.cell(row=desc_row, column=col_index, value=col_info['description']).border = thin_border
                    ws.cell(row=desc_row, column=col_index).alignment = left_align
                    # Fill enrichment method
                    ws.cell(row=enrich_row, column=col_index, value=col_info['enrichment']).border = thin_border
                    ws.cell(row=enrich_row, column=col_index).alignment = left_align

        # Add empty spacer row
        current_row += 1

    # Adjust column widths
    for col in range(1, 13):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 32

    # Set row heights for text wrapping
    for row in range(1, current_row + 1):
        ws.row_dimensions[row].height = 48

    wb.save('D:/rr_repo/manual/everything_columnmapping_temp.xlsx')
    print('Wrote to temp file: everything_columnmapping_temp.xlsx')
    print('Updated "Result I want" sheet - all used columns included (non-custom + custom).')
    print(f'Processed {len(asset_usage)} asset types.')


if __name__ == "__main__":
    main()
