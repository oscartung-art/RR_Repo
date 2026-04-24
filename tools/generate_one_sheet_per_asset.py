#!/usr/bin/env python3
"""
Generate Excel with one sheet per asset type, each sheet lists all columns used
"""

import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter


def main():
    # Full column metadata with descriptions from everything_columnmapping.md
    # All assets use custom_property_0 through custom_property_9, just with different meanings
    asset_descriptions = {
        'Furniture': {
            'custom_property_0': 'Primary asset classification',
            'custom_property_1': 'Model name or designer name',
            'custom_property_2': 'Brand/Designer/Collection identifier',
            'custom_property_3': 'Style or era classification',
            'custom_property_4': 'Primary color or material or surface finish',
            'custom_property_5': 'Usage context or location',
            'custom_property_6': 'Shape or physical configuration',
            'custom_property_7': 'Dimensions or scale classification',
            'custom_property_8': 'Reference code',
            'custom_property_9': 'Reserved',
        },
        'Fixture': {
            'custom_property_0': 'Primary asset classification',
            'custom_property_1': 'Model name or designer name',
            'custom_property_2': 'Brand/Designer/Collection identifier',
            'custom_property_3': 'Style or era classification',
            'custom_property_4': 'Primary color or material or surface finish',
            'custom_property_5': 'Usage context or location',
            'custom_property_6': 'Shape or physical configuration',
            'custom_property_7': 'Dimensions or scale classification',
            'custom_property_8': 'Reference code',
            'custom_property_9': 'Reserved',
        },
        'Vegetation': {
            'custom_property_0': 'Primary asset classification',
            'custom_property_1': 'Common Name of Vegetation',
            'custom_property_2': 'Plant Secondary Besides Green',
            'custom_property_3': 'Plant Shape',
            'custom_property_4': 'Chinese scientific name',
            'custom_property_5': 'Latin scientific name',
            'custom_property_6': 'Plant Height, Spread, Radius..etc',
            'custom_property_7': 'Reserved',
            'custom_property_8': 'Reference code',
            'custom_property_9': 'Reserved',
        },
        'Object': {
            'custom_property_0': 'Primary asset classification',
            'custom_property_1': 'Model name or designer name',
            'custom_property_2': 'Brand/Designer/Collection identifier',
            'custom_property_3': 'Style or era classification',
            'custom_property_4': 'Primary color or material or surface finish',
            'custom_property_5': 'Usage context or location',
            'custom_property_6': 'Shape or physical configuration',
            'custom_property_7': 'Dimensions or scale classification',
            'custom_property_8': 'Reference code',
            'custom_property_9': 'Reserved',
        },
        'Material': {
            'custom_property_0': 'Primary asset classification',
            'custom_property_1': 'Secondary Asset Classification',
            'custom_property_2': 'Specific Name/Model of Material',
            'custom_property_3': 'Style or era classification',
            'custom_property_4': 'Primary color',
            'custom_property_5': 'Reserved',
            'custom_property_6': 'Dimensions or scale classification',
            'custom_property_7': 'Reserved',
            'custom_property_8': 'Reference code',
            'custom_property_9': 'Reserved',
        },
        'Vehicle': {
            'custom_property_0': 'Primary asset classification',
            'custom_property_1': 'Model name or designer name',
            'custom_property_2': 'Brand/Designer/Collection identifier',
            'custom_property_3': 'Style or era classification',
            'custom_property_4': 'Primary color or material or surface finish',
            'custom_property_5': 'Usage context or location',
            'custom_property_6': 'Shape or physical configuration',
            'custom_property_7': 'Dimensions or scale classification',
            'custom_property_8': 'Reference code',
            'custom_property_9': 'Reserved',
        },
        'Layouts': {
            'custom_property_0': 'Primary asset classification',
            'custom_property_1': 'Asset inside the thumbnail',
            'custom_property_2': 'Usage context or location',
            'custom_property_3': 'Reserved',
            'custom_property_4': 'Reserved',
            'custom_property_5': 'Reserved',
            'custom_property_6': 'Reserved',
            'custom_property_7': 'Reserved',
            'custom_property_8': 'Reference code',
            'custom_property_9': 'Reserved',
        },
        'People': {
            'custom_property_0': 'Primary asset classification',
            'custom_property_1': 'Gender',
            'custom_property_2': 'Ethnicity',
            'custom_property_3': 'Usage context or location',
            'custom_property_4': 'Reserved',
            'custom_property_5': 'Reserved',
            'custom_property_6': 'Reserved',
            'custom_property_7': 'Reserved',
            'custom_property_8': 'Reference code',
            'custom_property_9': 'Reserved',
        },
        'VFX': {
            'custom_property_0': 'Primary asset classification',
            'custom_property_1': 'Reserved',
            'custom_property_2': 'Reserved',
            'custom_property_3': 'Reserved',
            'custom_property_4': 'Reserved',
            'custom_property_5': 'Reserved',
            'custom_property_6': 'Reserved',
            'custom_property_7': 'Reserved',
            'custom_property_8': 'Reserved',
            'custom_property_9': 'Reserved',
        },
        'Schedule': {
            'custom_property_0': 'Primary asset classification',
            'custom_property_1': 'Model name or designer name',
            'custom_property_2': 'Brand/Designer/Collection identifier',
            'custom_property_3': 'Style or era classification',
            'custom_property_4': 'Primary color or material or surface finish',
            'custom_property_5': 'Usage context or location',
            'custom_property_6': 'Shape or physical configuration',
            'custom_property_7': 'Dimensions or scale classification',
            'custom_property_8': 'Reference code',
            'custom_property_9': 'Reserved',
        },
    }

    # Create workbook
    wb = openpyxl.Workbook()

    # Remove default sheet
    default_sheet = wb.active
    wb.remove(default_sheet)

    # Styles
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_font_white = Font(bold=True, color='FFFFFF')
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)

    # Create sheet for each asset type
    for asset_name, descriptions in asset_descriptions.items():
        ws = wb.create_sheet(title=asset_name)

        # Headers
        headers = ['Custom Property', 'Description']
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font_white
            cell.fill = header_fill
            cell.border = thin_border
            cell.alignment = center_align

        # Fill data
        current_row = 2
        for cp_idx in range(10):
            cp_name = f'custom_property_{cp_idx}'
            desc = descriptions.get(cp_name, 'Reserved')
            ws.cell(row=current_row, column=1, value=cp_name).border = thin_border
            ws.cell(row=current_row, column=1).alignment = left_align

            ws.cell(row=current_row, column=2, value=desc).border = thin_border
            ws.cell(row=current_row, column=2).alignment = left_align

            current_row += 1

        # Column widths
        ws.column_dimensions['A'].width = 24
        ws.column_dimensions['B'].width = 45

        # Row heights
        for row in range(2, current_row):
            ws.row_dimensions[row].height = 40

    # Add overview sheet
    ws_overview = wb.create_sheet(title='Overview', index=0)
    headers = ['Asset Type', 'custom_property_0', 'custom_property_1', 'custom_property_2', 'custom_property_3', 'custom_property_4', 'custom_property_5', 'custom_property_6', 'custom_property_7', 'custom_property_8', 'custom_property_9']
    for col_idx, header in enumerate(headers, start=1):
        cell = ws_overview.cell(row=1, column=col_idx, value=header)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = center_align

    current_row = 2
    for asset_name, descriptions in asset_descriptions.items():
        ws_overview.cell(row=current_row, column=1, value=asset_name).border = thin_border
        ws_overview.cell(row=current_row, column=1).alignment = left_align
        ws_overview.cell(row=current_row, column=1).font = Font(bold=True)

        for cp_idx in range(10):
            cp_name = f'custom_property_{cp_idx}'
            desc = descriptions.get(cp_name, '-')
            col_idx = cp_idx + 2
            ws_overview.cell(row=current_row, column=col_idx, value=desc).border = thin_border
            ws_overview.cell(row=current_row, column=col_idx).alignment = left_align

        current_row += 1

    # Adjust overview column widths
    ws_overview.column_dimensions['A'].width = 18
    for i in range(2, 12):
        ws_overview.column_dimensions[get_column_letter(i)].width = 28

    # Set row heights for overview
    for row in ws_overview.iter_rows():
        ws_overview.row_dimensions[row[0].row].height = 45

    # Save
    output_path = 'D:/rr_repo/manual/everything_columnmapping_by_asset.xlsx'
    wb.save(output_path)
    print(f'Generated: {output_path}')
    print(f'Total sheets: {len(wb.sheetnames)}')


if __name__ == "__main__":
    main()
