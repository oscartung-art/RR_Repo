#!/usr/bin/env python3
"""
Generate Excel (xlsx) file from everything_columnmapping.md with asset-specific mappings
"""

import json
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter


def main():
    # First, let's use our existing column mapping data
    data_path = Path(r'manual/column_mapping_data.json')
    if not data_path.exists():
        print('column_mapping_data.json not found!')
        return

    with open(data_path, 'r', encoding='utf-8') as f:
        asset_data = json.load(f)

    # Create workbook
    wb = openpyxl.Workbook()

    # Styles
    header_font = Font(bold=True, size=12)
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_font_white = Font(bold=True, size=12, color='FFFFFF')
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin'),
    )

    # Create a sheet for each asset type
    for idx, (asset_type, mappings) in enumerate(asset_data.items()):
        # Create or get sheet
        if idx == 0:
            ws = wb.active
            ws.title = asset_type
        else:
            ws = wb.create_sheet(title=asset_type)

        # Set up headers
        # Get the first mapping to determine what columns we need
        if mappings:
            first_mapping = mappings[0]
            headers = list(first_mapping.keys())
        else:
            headers = []

        # Write headers
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font_white
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = thin_border

        # Write mappings
        for row_idx, mapping in enumerate(mappings, start=2):
            for col_idx, header in enumerate(headers, start=1):
                value = mapping.get(header, '')
                if value is None:
                    value = ''
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.alignment = left_align
                cell.border = thin_border

        # Adjust column widths
        for i in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(i)].width = 25

        # Set row heights for better readability
        for row in ws.iter_rows():
            ws.row_dimensions[row[0].row].height = 45

    # Create an overview sheet
    ws = wb.create_sheet(title='Overview', index=0)

    # Overview header
    overview_headers = ['Asset Type', 'custom_property_0', 'custom_property_1', 'custom_property_2', 'custom_property_3', 'custom_property_4', 'custom_property_5', 'custom_property_6', 'custom_property_7', 'custom_property_8', 'custom_property_9']

    for col_idx, header in enumerate(overview_headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border

    # Fill overview with descriptions instead of old semantic names
    row_idx = 2
    for asset_type, mappings in asset_data.items():
        ws.cell(row=row_idx, column=1, value=asset_type).font = Font(bold=True)
        ws.cell(row=row_idx, column=1).alignment = left_align
        ws.cell(row=row_idx, column=1).border = thin_border

        # For each custom property 0-9, get the description from the mappings
        for cp_idx in range(10):
            cp_name = f'custom_property_{cp_idx}'
            description = '-'
            # Find the mapping for this custom property
            for mapping in mappings:
                if mapping.get('Custom Property') == cp_name:
                    description = mapping.get('Description', '-')
                    break
            col_idx = cp_idx + 2
            cell = ws.cell(row=row_idx, column=col_idx, value=description)
            cell.alignment = left_align
            cell.border = thin_border

        row_idx += 1

    # Adjust overview column widths
    ws.column_dimensions['A'].width = 18
    for i in range(2, 12):
        ws.column_dimensions[get_column_letter(i)].width = 28

    # Set row heights for overview
    for row in ws.iter_rows():
        ws.row_dimensions[row[0].row].height = 45

    # Save the workbook
    output_path = Path(r'manual/everything_columnmapping.xlsx')
    wb.save(output_path)
    print(f'Generated: {output_path}')
    print(f'Sheets created: {len(wb.sheetnames)}')


if __name__ == "__main__":
    main()
